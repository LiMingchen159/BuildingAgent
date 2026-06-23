#!/usr/bin/env python3
"""Regenerate data/project_element/kb/KB_CATALOG_SUMMARY.md from brick_model.ttl + Plant API xlsx."""

from __future__ import annotations

import re
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    raise SystemExit("openpyxl required: pip install openpyxl")

ROOT = Path(__file__).resolve().parents[1]
KB = ROOT / "data" / "project_element" / "kb"
TTL_PATH = KB / "brick_model.ttl"
XLSX_PATH = KB / "Elements Chiller Plant API.xlsx"
OUT_PATH = KB / "KB_CATALOG_SUMMARY.md"

HL_SHEET = "DDC-GF-01(HL)"
PM_SHEETS = {"DDC-GF-01(PM)", "DDC-GF-01(MCC2PM)"}
WCC_REFERENCE = "WCC_01"
PLANT_CALC_SUFFIXES = ("_COP", "_DeltaT", "_P", "_Q")


def parse_ttl_points(text: str) -> list[dict]:
    points: list[dict] = []
    for block in text.split("\n\n"):
        if "brick:isPointOf" not in block:
            continue
        entity = re.search(r"^bldg:([^\s]+)\s+a\s+(?:brick:|bldg:)([^;\s]+)", block, re.M)
        parent = re.search(r"brick:isPointOf (bldg:[^\s;]+)", block)
        label = re.search(r'rdfs:label\s+"([^"]*)"', block)
        desc = re.search(r'bldg:sourceDescription\s+"([^"]*)"', block)
        sheet = re.search(r'bldg:sourceSheetName\s+"([^"]*)"', block)
        comment = re.search(r'rdfs:comment\s+"([^"]*)"', block)
        if not entity or not parent:
            continue
        points.append(
            {
                "id": entity.group(1),
                "brick": entity.group(2),
                "parent": parent.group(1).replace("bldg:", ""),
                "label": label.group(1) if label else "",
                "description": desc.group(1) if desc else "",
                "sheet": sheet.group(1) if sheet else "",
                "comment": comment.group(1) if comment else "",
            }
        )
    return points


def parse_equipment(text: str) -> list[tuple[str, str]]:
    pattern = re.compile(
        r"bldg:([^\s]+)\s+a\s+brick:(Water_Cooled_Chiller|Chilled_Water_Pump|Pump|Heat_Exchanger|HVAC_System|Loop|Ice_Storage|Cooling_Tower|Fan|Valve|Meter|Chiller)\s*;"
    )
    return [(m.group(1), m.group(2)) for m in pattern.finditer(text)]


def brick_type_stats(points: list[dict]) -> list[tuple[str, int]]:
    counts = Counter(p["brick"] for p in points)
    return sorted(counts.items(), key=lambda item: (-item[1], item[0]))


def xlsx_stats() -> dict:
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    sheets: dict[str, dict] = {}
    all_names: set[str] = set()
    for sheet_name in wb.sheetnames:
        if sheet_name == "Sheet1":
            continue
        ws = wb[sheet_name]
        names: set[str] = set()
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            if row[1]:
                name = str(row[1]).strip()
                names.add(name)
                all_names.add(name)
        sheets[sheet_name] = {"row_count": len(names), "unique_names": len(names)}
    wb.close()
    return {"sheets": sheets, "union_unique": len(all_names)}


def md_table(headers: list[str], rows: list[list[str]]) -> str:
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(cell.replace("|", "\\|") for cell in row) + " |")
    return "\n".join(lines)


def sort_points(points: list[dict]) -> list[dict]:
    return sorted(points, key=lambda p: (p["label"], p["id"]))


def wcc_hl_points(points: list[dict], parent: str = WCC_REFERENCE) -> list[dict]:
    return sort_points([p for p in points if p["parent"] == parent and p["sheet"] == HL_SHEET])


def wcc_pm_points(points: list[dict], parent: str = WCC_REFERENCE) -> list[dict]:
    return sort_points([p for p in points if p["parent"] == parent and p["sheet"] in PM_SHEETS])


def wcc_hydraulic_points(points: list[dict], parent: str = WCC_REFERENCE) -> list[dict]:
    return sort_points(
        [p for p in points if p["parent"] == parent and p["sheet"] not in PM_SHEETS and p["sheet"] != HL_SHEET]
    )


def parent_points(points: list[dict], parent: str) -> list[dict]:
    return sort_points([p for p in points if p["parent"] == parent])


def wcc_index(parent: str) -> int:
    return int(parent.split("_")[1])


def hl_suffix(label: str, n: int = 1) -> str:
    prefix = f"WCC_{n}_"
    return label[len(prefix) :] if label.startswith(prefix) else label


def plant_suffix(label: str, n: int = 1) -> str:
    prefix = f"WCC-L1-{n:02d}"
    if label.startswith(prefix):
        return label[len(prefix) :]
    return label


def point_rows(points: list[dict], include_sheet: bool = True) -> list[list[str]]:
    rows: list[list[str]] = []
    for p in points:
        row = [p["label"] or p["id"], p["brick"], p["description"] or "—"]
        if include_sheet:
            row.append(p["sheet"] or "—")
        rows.append(row)
    return rows


def point_rows_with_parent(points: list[dict]) -> list[list[str]]:
    rows: list[list[str]] = []
    for p in points:
        rows.append(
            [
                p["parent"],
                p["label"] or p["id"],
                p["brick"],
                p["description"] or "—",
                p["sheet"] or "—",
            ]
        )
    return rows


def equipment_section(points: list[dict], parent: str, title: str, note: str) -> str:
    pts = parent_points(points, parent)
    headers = ["BMS name / label", "Brick type", "Description", "sourceSheet"]
    body = md_table(headers, point_rows(pts))
    return f"### {title}\n\n**{len(pts)}** points. {note}\n\n{body}\n"


def suffix_reference_table(
    points: list[dict],
    suffix_fn,
    n: int,
    headers: list[str],
) -> str:
    rows = []
    for p in points:
        suffix = suffix_fn(p["label"], n)
        rows.append([suffix, p["brick"], p["description"] or "—", p["sheet"] or "—"])
    return md_table(headers, rows)


def layer_for_point(point: dict) -> str:
    if point["sheet"] == HL_SHEET:
        return "HL"
    if point["sheet"] in PM_SHEETS:
        return "PM"
    return "Plant"


POC_SECTION = """
## 4. POC project PDFs (`Poc Project-20260517T050008Z-3-001/Poc Project/`)

Hong Kong **ELEMENTS** mall MEP document pack, grouped by subsystem. Descriptions below are **inferred from filenames** (not full OCR).

### 4.1 Chiller system (4 files)

| File | Likely content |
|------|----------------|
| `Equipment Schedule_Chiller.pdf` | Chiller equipment schedule: model, capacity, qty, location |
| `Equipment Schedule_Chilled Water Pump.pdf` | Chilled-water pump schedule |
| `Equipment Schedule_Condenser Water Pump.pdf` | Condenser-water pump schedule |
| `Chiller Control Panel Line Diagram.pdf` | Chiller plant control panel / logic diagram |

### 4.2 Sea-water cooling system (22 files)

Sea-water cooling, plate heat exchangers, plant room layouts — related to SWP BMS points.

| File | Likely content |
|------|----------------|
| `Overall Equipment Schedule_Sea Water System.pdf` | Sea-water system equipment schedule |
| `Overall Equipment Specification_Sea Water System.pdf` | Sea-water equipment specifications |
| `Seawater System Scenario.pdf` | Operating scenarios / modes |
| `(Schematic) Overall Schematic Diagram_Sea Water System.pdf` | Overall sea-water schematic |
| `(Schematic) Sea Water Plant Room (1).pdf` / `(2).pdf` | Sea-water pump room schematics |
| `(Schematic) Chiller Plant Room.pdf` | Chiller plant room schematic |
| `(Layout) Underground Plant Room.pdf` | Underground plant room layout |
| `(Layout) Chiller Plant Room.pdf` | Chiller plant room layout |
| `(Layout) Heat Exchanger Room (1).pdf` / `(2).pdf` | Heat exchanger room layouts |
| `(Layout) Typical Sea Water Pump.pdf` | Typical sea-water pump layout |
| `(Control Philosophy) main AC plant.pdf` | Main AC plant control philosophy |
| `(Control Philosophy) Seawater Cooling Plant.pdf` | Sea-water plant control philosophy |
| `(As-fit Drawing) 03-14-2008_Seawater_Cooling_Plant.pdf` | 2008 as-built sea-water plant |
| `(O&M) Auto Backwash Strainer.pdf` | Auto backwash strainer O&M |
| `【1】O&M Manual/Sea Water Plant - OnM Manual Vol. 1~3.pdf` | Sea-water plant O&M manual (3 vols) |
| `【2】Heat Exchanger/Equipment - Plate Heat Exchanger.pdf` | Plate heat exchanger equipment data |
| `【2】Heat Exchanger/Heat Exchanger_Product Manual.pdf` | Heat exchanger product manual |
| `【2】Heat Exchanger/Heat Exchanger_Name Plate.pdf` | Heat exchanger nameplate |

### 4.3 AHU, PAU & FCU — air side (10 files)

Air side: AHUs, PAUs, fan coil units.

| File | Likely content |
|------|----------------|
| `Equipment Schedule_AHU & PAU.pdf` | AHU / PAU equipment schedule |
| `Equipment Schedule_Fan Coil Unit.pdf` | FCU equipment schedule |
| `(As-fitted) MVAC-01 V4.pdf` / `MVAC-02.pdf` | As-fitted MVAC drawings 01 / 02 |
| `【Air Side】Mall Schematic/2373-M-S06_09-Layout1_S06~S09.pdf` | Mall duct schematic zones S06–S09 |
| `【Air Side】Mall Schematic/2373-M-S10-Layout1.pdf` / `S11-Layout1.pdf` | Mall duct schematic S10 / S11 |

### 4.4 ELEMENTS House Rule & Safety Control (6 files)

Property safety and work-permit rules — **not BMS data**; use for compliance / site access.

| File | Likely content |
|------|----------------|
| `App A - PS- EleRules.pdf` | MEP rules |
| `App B - Permit-to-Work Guideline (2025).pdf` | Permit-to-work guideline |
| `App C - Risk Assessment and Control.pdf` | Risk assessment and control |
| `App D - PS - Safety and Health Rules.pdf` | Safety and health rules |
| `MMRW_5.0_20240501R_.pdf` | MEP maintenance regulations |
| `SafM_6.0_20250101_.pdf` | Safety manual |

### 4.5 MVAC & Sea Water Plant O&M (4 files)

| File | Likely content |
|------|----------------|
| `O&M Manual for MVAC system.pdf` | Building-wide MVAC O&M manual (large) |
| `Sea Water Plant - OnM Manual Vol. 1~3.pdf` | Sea-water O&M (may overlap §4.2; use newer copy) |
""".strip()


def generate() -> str:
    text = TTL_PATH.read_text(encoding="utf-8", errors="ignore")
    points = parse_ttl_points(text)
    equipment = parse_equipment(text)
    xlsx = xlsx_stats()

    entity_count = len(re.findall(r"^bldg:[^\s]+\s+a\s+(?:brick:|bldg:)", text, re.M))
    parents = sorted({p["parent"] for p in points})
    by_parent = {parent: parent_points(points, parent) for parent in parents}

    wcc_parents = sorted([p for p in parents if re.fullmatch(r"WCC_\d{2}", p)], key=wcc_index)
    chp_parents = sorted(p for p in parents if p.startswith("CHP_"))
    swp_parents = sorted(p for p in parents if p.startswith("SWP_"))

    hl_ref = wcc_hl_points(points)
    pm_ref = wcc_pm_points(points)
    hyd_ref = wcc_hydraulic_points(points)

    hl_per = len(hl_ref)
    pm_per = len(pm_ref)
    hyd_per = len(hyd_ref)
    per_chiller_total = hl_per + pm_per + hyd_per
    assert hl_per > 0 and pm_per > 0 and hyd_per > 0
    assert len(wcc_parents) == 8
    for parent in wcc_parents:
        assert len(wcc_hl_points(points, parent)) == hl_per, parent
        assert len(wcc_pm_points(points, parent)) == pm_per, parent
        assert len(wcc_hydraulic_points(points, parent)) == hyd_per, parent

    hl_total = hl_per * len(wcc_parents)
    pm_total = pm_per * len(wcc_parents)
    hyd_total = hyd_per * len(wcc_parents)
    wcc_points_total = per_chiller_total * len(wcc_parents)

    cop_points = sort_points([p for p in points if p["label"].endswith("_COP") and p["parent"] in wcc_parents])
    assert len(cop_points) == 8

    def is_plant_calc_point(point: dict) -> bool:
        label = point["label"]
        return label.startswith("WCC-L1-") and any(label.endswith(suffix) for suffix in PLANT_CALC_SUFFIXES)

    plant_calc_points = sort_points(
        [p for p in points if p["parent"] in wcc_parents and is_plant_calc_point(p)]
    )
    assert len(plant_calc_points) == 32

    all_plant_hyd = sort_points(
        [p for p in points if p["parent"] in wcc_parents and layer_for_point(p) == "Plant"]
    )

    per_chiller_rows = []
    for parent in wcc_parents:
        n = wcc_index(parent)
        per_chiller_rows.append(
            [
                f"`WCC_{n}` / `{parent}`",
                str(hl_per),
                str(pm_per),
                str(hyd_per),
                str(per_chiller_total),
                f"`WCC_{n}_*`",
                f"`WCC-L1-{n:02d}*`",
            ]
        )

    equip_lines = []
    equip_counter = Counter(kind for _, kind in equipment)
    for kind, count in sorted(equip_counter.items()):
        equip_lines.append(f"- `{kind}` × {count}")

    equipment_count_rows = []
    for parent in wcc_parents + chp_parents + swp_parents:
        equipment_count_rows.append([parent, str(len(by_parent[parent]))])
    if "Ice_Rink_Chiller_Plant" in by_parent:
        equipment_count_rows.append(["Ice_Rink_Chiller_Plant", str(len(by_parent["Ice_Rink_Chiller_Plant"]))])
    for parent in sorted(
        p
        for p in parents
        if p not in wcc_parents and p not in chp_parents and p not in swp_parents and p != "Ice_Rink_Chiller_Plant"
    ):
        equipment_count_rows.append([parent, str(len(by_parent[parent]))])

    xlsx_sheet_lines = [
        [f"`{name}`", str(info["unique_names"])]
        for name, info in sorted(xlsx["sheets"].items())
    ]

    stats = brick_type_stats(points)
    stats_rows = [[brick, str(count)] for brick, count in stats[:30]]
    if len(stats) > 30:
        stats_rows.append(["…", f"+{len(stats) - 30} more types"])

    special_index_rows = [
        [
            "COP (coefficient of performance)",
            "`WCC-L1-01_COP` … `WCC-L1-08_COP`",
            "Plant",
            "Yes (not in HL)",
            "`bms_points_query(q=\"COP\")` or §2.3.1",
        ],
        [
            "Plant cooling / power / delta-T",
            "`WCC-L1-0n_DeltaT` / `_P` / `_Q`",
            "Plant",
            "Yes",
            "§2.3.2",
        ],
        [
            "Live motor power",
            "`WCC_{1-8}_TLKW`",
            "HL",
            "No",
            "`bms_live_read` / `?q=TLKW`",
        ],
        [
            "Leaving chilled-water temp",
            "`WCC_{1-8}_SUWT`",
            "HL",
            "No",
            "§2.1",
        ],
        [
            "MCC amps / energy",
            "`GF_2000A_WCC_L1_01_*` etc.",
            "PM",
            "Yes",
            "§2.2",
        ],
        [
            "Chilled-water pumps",
            "`CHP-1P-D01-S` etc.",
            "CHP",
            "—",
            "§2.4",
        ],
        [
            "Sea-water pumps",
            "`SWP_01_*` etc.",
            "SWP",
            "—",
            "§2.5",
        ],
    ]

    today = date.today().isoformat()
    parts = [
        "# Element KB Catalog Summary",
        "",
        f"> **Path:** `data/project_element/kb/`  ",
        f"> **Updated:** {today} (`scripts/regenerate_kb_catalog_summary.py`)  ",
        "> **Purpose:** Quick index of KB files and BMS point layers — read before blind exploration.  ",
        "> **Agent:** Read §1 + §5 first; use HL / PM / Plant naming layers. **Do not** search only `WCC_{n}_*` HL for COP and other Plant points.",
        "",
        "---",
        "",
        "## 1. KB overview",
        "",
        md_table(
            ["Category", "Path", "Count", "Notes"],
            [
                [
                    "Brick semantic model",
                    "`brick_model.ttl`",
                    f"{entity_count} entities / {len(points)} point relations",
                    "Full plant: chillers, pumps, valves, ice rink; `urn:hensen_chiller_plant#`",
                ],
                [
                    "Plant API point list",
                    "`Elements Chiller Plant API.xlsx`",
                    f"{xlsx['union_unique']} unique names / {len(xlsx['sheets'])} sheets",
                    "enteliWEB export; HL / PM / MCC / sea-water plant",
                ],
                [
                    "BMS data guide",
                    "`bms_guide.md`",
                    "1 file",
                    "Agent quick ref + live/history APIs + tools",
                ],
                [
                    "POC drawings & manuals",
                    "`Poc Project-20260517T050008Z-3-001/`",
                    "46 PDFs",
                    "ELEMENTS mall MVAC / sea-water cooling / chillers / safety rules",
                ],
            ],
        ),
        "",
        "### 1.1 Three-layer chiller point model (read first)",
        "",
        f"Each chiller **{per_chiller_total}** points = **{hl_per} HL** + **{pm_per} PM** + **{hyd_per} Plant**. Plant total **{wcc_points_total}** (8×{per_chiller_total}).",
        "",
        md_table(
            ["Layer", "Name pattern", "Source sheet", "Per chiller", "Plant total", "Notes"],
            [
                [
                    "**HL (high-level)**",
                    "`WCC_{1-8}_TLKW`, `WCC_3_Run_Status`",
                    "`DDC-GF-01(HL)`",
                    str(hl_per),
                    str(hl_total),
                    "Chiller controller HL interface; default for live ops queries",
                ],
                [
                    "**PM (power)**",
                    "`GF_2000A_WCC_L1_01_AMP_L1` etc.",
                    "`DDC-GF-01(PM)` / `(MCC2PM)`",
                    str(pm_per),
                    str(pm_total),
                    "MCC metering; L1 unit numbering",
                ],
                [
                    "**Plant (room DDC)**",
                    "`WCC-L1-01-CHWST`, `WCC-L1-03_COP`",
                    "`DDC-GF-01` / `02` / `03`",
                    str(hyd_per),
                    str(hyd_total),
                    "Dry-contact DDC hydraulics/status; **COP lives here, not HL**",
                ],
            ],
        ),
        "",
        "### 1.2 Per-chiller scale",
        "",
        md_table(
            ["Chiller", "HL", "PM", "Plant", "Total", "HL prefix", "Plant prefix"],
            per_chiller_rows,
        ),
        "",
        "### 1.3 Special points index (easy to miss)",
        "",
        "These **cannot** be found with HL-only `WCC_{n}_*` search patterns.",
        "",
        md_table(
            ["Topic", "Name pattern", "Layer", "Not in HL?", "How to find"],
            special_index_rows,
        ),
        "",
        "**Data links:**",
        "",
        f"- HL ({hl_total}) maps 1:1 to BMS-database `points.name` and enteliWEB.",
        "- Plant layer has **8 direct COP points** (`WCC-L1-0n_COP`) plus `_DeltaT` / `_P` / `_Q`.",
        "- CHP / SWP / ice rink have **no** COP points.",
        "- **Values:** >3 points / history / batch → local BMS-database (`bms_points_query` / `bms_timeseries_query` per `skill_element_bms_data`); ≤3 live/alarm → `bms_live_read`. Names/catalog → this doc or `bms_points_query` once.",
        "",
        "**Excel sheets:**",
        "",
        md_table(["Sheet", "Unique names"], xlsx_sheet_lines),
        "",
        "---",
        "",
        "## 2. Full-plant Brick model (`brick_model.ttl`)",
        "",
        "| Item | Value |",
        "|------|-------|",
        "| Format | Turtle (`.ttl`) |",
        "| Namespace | `bldg:` = `urn:hensen_chiller_plant#` |",
        f"| Entities | **{entity_count}** |",
        f"| Point relations | **{len(points)}** `brick:isPointOf` |",
        f"| Chiller HL | **8 × {hl_per} = {hl_total}** |",
        f"| Chiller PM | **8 × {pm_per} = {pm_total}** |",
        f"| Chiller Plant | **8 × {hyd_per} = {hyd_total}** |",
        "",
        "**Equipment entities:**",
        "",
        "\n".join(equip_lines),
        "",
        "**Point counts per parent:**",
        "",
        md_table(["Equipment / parent", "Points"], equipment_count_rows),
        "",
        "**BMS naming (three layers — do not mix):**",
        "",
        "| Layer | Brick equipment ID | BMS name example | Agent note |",
        "|-------|-------------------|------------------|------------|",
        "| HL | `WCC_01` … `WCC_08` | `WCC_1_TLKW` (single digit 1–8) | Run/temp/power; **no COP** |",
        "| PM | same | `GF_2000A_WCC_L1_01_KWH` | MCC / L1 unit numbering |",
        "| Plant | same | `WCC-L1-03_COP`, `WCC-L1-01-CHWST` | Room DDC; COP / flow / valves |",
        "",
        "- Do not use `WCC_01_TLKW` (two-digit Brick ID ≠ BMS name).",
        "- Do not search `WCC_3_COP` for COP — correct name is `WCC-L1-03_COP`.",
        "",
        "---",
        "",
        f"## 2.1 Chiller HL points ({hl_per} per unit, full list)",
        "",
        f"All **{hl_per}** High-Level points for **WCC_1** (`DDC-GF-01(HL)`).",
        f"For **WCC_2** … **WCC_8**, replace `WCC_1` with `WCC_n` (plant total **{hl_total}**). **HL has no COP.**",
        "",
        md_table(
            ["BMS name", "Brick type", "Description", "sourceSheet"],
            point_rows(hl_ref),
        ),
        "",
        "### 2.1.1 HL suffix quick ref (WCC_1 template)",
        "",
        "Prefix each suffix with `WCC_{n}_` for chiller *n*.",
        "",
        suffix_reference_table(
            hl_ref,
            hl_suffix,
            1,
            ["HL suffix", "Brick type", "Description", "sourceSheet"],
        ),
        "",
        "---",
        "",
        f"## 2.2 Chiller PM / MCC power points ({pm_per} per unit)",
        "",
        f"All **{pm_per}** PM / MCC points for **WCC_01**. Other chillers are isomorphic by L1 unit number (plant total **{pm_total}**).",
        "",
        md_table(
            ["BMS name / label", "Brick type", "Description", "sourceSheet"],
            point_rows(pm_ref),
        ),
        "",
        "### 2.2.1 PM suffix quick ref (WCC_01 / L1-01 template)",
        "",
        suffix_reference_table(
            pm_ref,
            lambda label, _n: label.replace("GF_2000A_WCC_L1_01", "GF_2000A_WCC_L1_NN"),
            1,
            ["PM name pattern (NN = unit no.)", "Brick type", "Description", "sourceSheet"],
        ),
        "",
        "---",
        "",
        "## 2.3 Chiller Plant (room DDC) points",
        "",
        "Plant points come from dry-contact DDC (`DDC-GF-01` / `02` / `03`), **not** `DDC-GF-01(HL)`.",
        "",
        "### 2.3.1 Plant COP points (8, full list)",
        "",
        "One direct COP reading per chiller; Brick type `Coefficient_Of_Performance_Sensor` (custom `bldg:` namespace).",
        "",
        md_table(
            ["Chiller", "BMS name", "Brick type", "Description", "sourceSheet", "Semantic note"],
            [
                [
                    f"WCC_{wcc_index(p['parent']):02d}",
                    p["label"],
                    p["brick"],
                    p["description"] or "—",
                    p["sheet"],
                    (p["comment"][:80] + "…") if len(p["comment"]) > 80 else (p["comment"] or "coefficient of performance"),
                ]
                for p in cop_points
            ],
        ),
        "",
        "**BMS search:** `GET /api/v1/points?q=COP`",
        "",
        "### 2.3.2 Plant calc-related points (COP / DeltaT / P / Q — 32 total)",
        "",
        md_table(
            ["Chiller", "BMS name", "Brick type", "Description", "sourceSheet"],
            point_rows_with_parent(plant_calc_points),
        ),
        "",
        f"### 2.3.3 Plant suffix quick ref (WCC-L1-01 template, {hyd_per} suffixes)",
        "",
        "Replace `WCC-L1-01` with `WCC-L1-0n` (n = 01…08) for each chiller's Plant names.",
        "",
        suffix_reference_table(
            hyd_ref,
            plant_suffix,
            1,
            ["Plant suffix", "Brick type", "Description", "sourceSheet"],
        ),
        "",
        f"### 2.3.4 WCC_01 Plant full list ({hyd_per} points, example)",
        "",
        md_table(
            ["BMS name / label", "Brick type", "Description", "sourceSheet"],
            point_rows(hyd_ref),
        ),
        "",
        f"### 2.3.5 Full-plant Plant hydraulic points (8×{hyd_per} = {hyd_total})",
        "",
        md_table(
            ["Chiller", "BMS name", "Brick type", "Description", "sourceSheet"],
            point_rows_with_parent(all_plant_hyd),
        ),
        "",
        "---",
        "",
        "## 2.4 Chilled-water pumps CHP (38 per unit × 10)",
        "",
        equipment_section(
            points,
            "CHP_1P_01",
            "CHP_1P_01 — full point list",
            "`CHP_1P_02` … `CHP_1P_10` each have **38** isomorphic points.",
        ),
        "",
        md_table(
            ["CHP unit", "Points"],
            [[parent, str(len(by_parent[parent]))] for parent in chp_parents],
        ),
        "",
        "---",
        "",
        "## 2.5 Sea-water pumps SWP, ice rink & system-level",
        "",
    ]

    for swp in swp_parents:
        parts.append(
            equipment_section(
                points,
                swp,
                f"{swp} — full point list ({len(by_parent[swp])} points)",
                "Sea-water pump power and status; **no COP**.",
            )
        )
        parts.append("")

    if "Ice_Rink_Chiller_Plant" in by_parent:
        parts.append(
            equipment_section(
                points,
                "Ice_Rink_Chiller_Plant",
                "Ice_Rink_Chiller_Plant — full point list",
                "Ice-rink chiller plant; **no COP**.",
            )
        )
        parts.append("")

    system_parents = [
        p
        for p in parents
        if p not in wcc_parents and p not in chp_parents and p not in swp_parents and p != "Ice_Rink_Chiller_Plant"
    ]
    for parent in sorted(system_parents):
        parts.append(
            equipment_section(points, parent, f"{parent} — full point list", "")
        )
        parts.append("")

    parts.extend(
        [
            "---",
            "",
            "## 2.6 Brick type stats (top 30)",
            "",
            md_table(["Brick type", "Count"], stats_rows),
            "",
            "---",
            "",
            "## 3. BMS integration docs",
            "",
            md_table(
                ["File", "Contents"],
                [
                    [
                        "`bms_guide.md`",
                        "§0 Agent quick ref；§1–4 live/history APIs & tools；Appendix A timeseries migration",
                    ],
                ],
            ),
            "",
            "---",
            "",
            POC_SECTION,
            "",
            "---",
            "",
            "## 5. Task → file routing",
            "",
            md_table(
                ["Task", "Open first"],
                [
                    ["KB structure & naming", "This doc §1.1–§1.3"],
                    [f"WCC HL points ({hl_per}/unit)", "§2.1 or §2.1.1"],
                    ["Chiller COP (direct points)", "§2.3.1; API `?q=COP`; **do not** search HL for `WCC_n_COP`"],
                    ["Plant flow / pressure / valve status", "§2.3.3–§2.3.5"],
                    ["Chiller PM / MCC power", "§2.2–§2.2.1"],
                    ["CHP / SWP point names", "§2.4–§2.5"],
                    ["Point history / trend", "`bms_guide.md` §3 → `/api/v1/timeseries`"],
                    ["Live chiller temp / power (≤3 points)", "`bms_live_read`; >3 or batch → `bms_points_query` local DB first"],
                    ["Chiller / pump equipment specs", "`Chiller System_* / Equipment Schedule_*.pdf`"],
                    ["Sea-water system design", "`Sea Water Cooling System_*`"],
                    ["Site access / compliance", "`ELEMENTS House Rule & Safety Control/*`"],
                ],
            ),
            "",
            "---",
            "",
            "## 6. Maintenance",
            "",
            "- After updating `brick_model.ttl` or `Elements Chiller Plant API.xlsx`, run:",
            "  ```bash",
            "  python3 scripts/regenerate_kb_catalog_summary.py",
            "  ```",
            "- New PDFs: add under POC subfolders, then manually update §4.",
            "- Removed redundancies: `Element Chiller High Level API.xlsx`, original POC zip.",
            "",
            "---",
            "",
            f"*Auto-generated by `scripts/regenerate_kb_catalog_summary.py` from `brick_model.ttl` and `Elements Chiller Plant API.xlsx` ({today}).*",
            "",
        ]
    )

    return "\n".join(parts)


def main() -> None:
    content = generate()
    OUT_PATH.write_text(content, encoding="utf-8")
    line_count = len(content.splitlines())
    print(f"Wrote {OUT_PATH} ({line_count} lines)")
    assert "WCC-L1-08_COP" in content
    assert "2.3.5" in content
    assert "Coefficient_Of_Performance_Sensor" in content


if __name__ == "__main__":
    main()
